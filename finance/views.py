from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.tokens import default_token_generator
from django.contrib import messages
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.db.models import Sum, Count, Avg, Q
from django.http import JsonResponse, HttpResponse
from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.utils import timezone
from django.conf import settings
from django.template.loader import render_to_string

import calendar
import json
from datetime import date, timedelta
from decimal import Decimal

from .models import Category, SubCategory, Transaction
from .forms import (
    LoginForm, PasswordResetRequestForm, CustomSetPasswordForm,
    CategoryForm, SubCategoryForm, TransactionForm, TransactionFilterForm
)


# ─── Auth Views ────────────────────────────────────────────────────────────────

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    error = None
    if request.method == 'POST':
        identifier = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')

        # Try by username first
        user = authenticate(request, username=identifier, password=password)

        # If failed, try finding by email then auth with their username
        if user is None:
            try:
                u = User.objects.get(email=identifier)
                user = authenticate(request, username=u.username, password=password)
            except User.DoesNotExist:
                pass

        if user is not None:
            login(request, user)
            messages.success(request, f'Welcome back, {user.get_full_name() or user.username}!')
            return redirect('dashboard')
        else:
            error = 'Invalid username/email or password. Please try again.'

    features = [
        'Track income and expenses',
        'Monthly reports and analytics',
        'Calendar overview',
        'Excel and PDF exports',
    ]
    return render(request, 'finance/login.html', {'error': error, 'features': features})


def logout_view(request):
    logout(request)
    messages.info(request, 'You have been logged out.')
    return redirect('login')


def password_reset_request(request):
    form = PasswordResetRequestForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        email = form.cleaned_data['email']
        user = User.objects.get(email=email)
        token = default_token_generator.make_token(user)
        uid = urlsafe_base64_encode(force_bytes(user.pk))
        reset_link = request.build_absolute_uri(f'/reset-password/{uid}/{token}/')
        subject = 'FaithLedger - Password Reset'
        message = f"""
Hello {user.get_full_name() or user.username},

You requested a password reset for your FaithLedger account.

Click the link below to reset your password (valid for 30 minutes):
{reset_link}

If you did not request this, please ignore this email.

Best regards,
FaithLedger Team
"""
        try:
            send_mail(subject, message, settings.DEFAULT_FROM_EMAIL, [email])
            messages.success(request, 'Password reset link sent to your Gmail. Please check your inbox.')
        except Exception:
            messages.error(request, 'Failed to send email. Please check your email configuration.')
        return redirect('login')
    return render(request, 'finance/password_reset_request.html', {'form': form})


def password_reset_confirm(request, uidb64, token):
    try:
        uid = force_str(urlsafe_base64_decode(uidb64))
        user = User.objects.get(pk=uid)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        user = None

    if user is None or not default_token_generator.check_token(user, token):
        messages.error(request, 'Password reset link is invalid or has expired.')
        return redirect('login')

    form = CustomSetPasswordForm(user, request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Password reset successful. You can now login.')
        return redirect('login')
    return render(request, 'finance/password_reset_confirm.html', {'form': form})


# ─── Dashboard ─────────────────────────────────────────────────────────────────
class DecimalEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        if isinstance(obj, date):
            return obj.isoformat()
        return super().default(obj)

@login_required
def dashboard(request):
    today = date.today()
    this_month = today.replace(day=1)
    next_month = (this_month.replace(day=28) + timedelta(days=4)).replace(day=1)

    # Monthly totals
    monthly_qs = Transaction.objects.filter(date__gte=this_month, date__lt=next_month)
    total_income = monthly_qs.filter(transaction_type='Income').aggregate(s=Sum('amount'))['s'] or Decimal('0')
    total_expense = monthly_qs.filter(transaction_type='Expense').aggregate(s=Sum('amount'))['s'] or Decimal('0')
    net_balance = total_income - total_expense
    pending_count = Transaction.objects.filter(status='Pending').count()

    # Last 10 transactions
    recent_transactions = Transaction.objects.select_related('category', 'subcategory').order_by('-date', '-created_at')[:10]

    # Monthly Income vs Expense (last 6 months)
    monthly_data = []
    for i in range(5, -1, -1):
        d = today.replace(day=1) - timedelta(days=i * 28)
        month_start = d.replace(day=1)
        month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
        inc = Transaction.objects.filter(date__gte=month_start, date__lt=month_end, transaction_type='Income').aggregate(s=Sum('amount'))['s'] or 0
        exp = Transaction.objects.filter(date__gte=month_start, date__lt=month_end, transaction_type='Expense').aggregate(s=Sum('amount'))['s'] or 0
        monthly_data.append({
            'month': month_start.strftime('%b %Y'),
            'income': float(inc),
            'expense': float(exp),
        })

    # Category-wise expense pie (this month)
    cat_expense = monthly_qs.filter(transaction_type='Expense').values('category__name').annotate(total=Sum('amount')).order_by('-total')

    # Sunday offering trend (last 6 months) - simplified as monthly offering totals
    offering_data = []
    try:
        sunday_cat = Category.objects.get(name='Sunday Offerings', type='Income')
        for item in monthly_data:
            month_start = date(*[int(x) for x in (item['month'].split(' ')[1], list(calendar.month_abbr).index(item['month'].split(' ')[0]), 1)])
            month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
            amt = Transaction.objects.filter(
                date__gte=month_start, date__lt=month_end,
                category=sunday_cat
            ).aggregate(s=Sum('amount'))['s'] or 0
            offering_data.append({'month': item['month'], 'amount': float(amt)})
    except (Category.DoesNotExist, Exception):
        offering_data = [{'month': m['month'], 'amount': 0} for m in monthly_data]

    context = {
        'total_income': total_income,
        'total_expense': total_expense,
        'net_balance': net_balance,
        'pending_count': pending_count,
        'recent_transactions': recent_transactions,
        'monthly_data': json.dumps(monthly_data),
        'cat_expense': json.dumps(list(cat_expense.values('category__name', 'total')), cls = DecimalEncoder),
        'offering_data': json.dumps(offering_data),
    }
    return render(request, 'finance/dashboard.html', context)


# ─── Transactions ──────────────────────────────────────────────────────────────

@login_required
def transaction_list(request):
    qs = Transaction.objects.select_related('category', 'subcategory', 'user').all()
    form = TransactionFilterForm(request.GET or None)

    if form.is_valid():
        if form.cleaned_data.get('date_from'):
            qs = qs.filter(date__gte=form.cleaned_data['date_from'])
        if form.cleaned_data.get('date_to'):
            qs = qs.filter(date__lte=form.cleaned_data['date_to'])
        if form.cleaned_data.get('transaction_type'):
            qs = qs.filter(transaction_type=form.cleaned_data['transaction_type'])
        if form.cleaned_data.get('category'):
            qs = qs.filter(category=form.cleaned_data['category'])
        if form.cleaned_data.get('status'):
            qs = qs.filter(status=form.cleaned_data['status'])
        if form.cleaned_data.get('search'):
            qs = qs.filter(notes__icontains=form.cleaned_data['search'])

    total_income = qs.filter(transaction_type='Income').aggregate(s=Sum('amount'))['s'] or 0
    total_expense = qs.filter(transaction_type='Expense').aggregate(s=Sum('amount'))['s'] or 0

    paginator = Paginator(qs, 20)
    page = request.GET.get('page', 1)
    transactions = paginator.get_page(page)

    return render(request, 'finance/transaction_list.html', {
        'transactions': transactions,
        'form': form,
        'total_income': total_income,
        'total_expense': total_expense,
    })


@login_required
def transaction_add(request):
    form = TransactionForm(request.POST or None, initial={'date': date.today()})
    if request.method == 'POST' and form.is_valid():
        txn = form.save(commit=False)
        txn.user = request.user
        txn.save()
        messages.success(request, 'Transaction added successfully!')
        return redirect('transaction_list')
    return render(request, 'finance/transaction_form.html', {'form': form, 'title': 'Add Transaction'})


@login_required
def transaction_edit(request, pk):
    txn = get_object_or_404(Transaction, pk=pk)
    form = TransactionForm(request.POST or None, instance=txn)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Transaction updated successfully!')
        return redirect('transaction_list')
    return render(request, 'finance/transaction_form.html', {'form': form, 'title': 'Edit Transaction', 'txn': txn})


@login_required
def transaction_delete(request, pk):
    txn = get_object_or_404(Transaction, pk=pk)
    if request.method == 'POST':
        txn.delete()
        messages.success(request, 'Transaction deleted successfully!')
        return redirect('transaction_list')
    return render(request, 'finance/transaction_confirm_delete.html', {'txn': txn})


# ─── AJAX ──────────────────────────────────────────────────────────────────────

@login_required
def get_categories(request):
    t_type = request.GET.get('type', '')
    cats = Category.objects.filter(type=t_type, is_active=True).values('id', 'name')
    return JsonResponse(list(cats), safe=False)


@login_required
def get_subcategories(request):
    cat_id = request.GET.get('category_id', '')
    subs = SubCategory.objects.filter(category_id=cat_id, is_active=True).values('id', 'name')
    return JsonResponse(list(subs), safe=False)


# ─── Calendar ──────────────────────────────────────────────────────────────────

@login_required
def calendar_view(request):
    today = date.today()
    year = int(request.GET.get('year', today.year))
    month = int(request.GET.get('month', today.month))

    # Navigate
    prev_month = date(year, month, 1) - timedelta(days=1)
    next_month_d = date(year, month, 28) + timedelta(days=4)
    next_month_d = next_month_d.replace(day=1)

    # Get all transactions for the month
    month_start = date(year, month, 1)
    month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    txns = Transaction.objects.filter(date__gte=month_start, date__lt=month_end)

    # Build daily summary dict
    daily = {}
    for t in txns:
        d = t.date.day
        if d not in daily:
            daily[d] = {'income': Decimal('0'), 'expense': Decimal('0'), 'pending': 0}
        if t.transaction_type == 'Income':
            daily[d]['income'] += t.amount
        else:
            daily[d]['expense'] += t.amount
        if t.status == 'Pending':
            daily[d]['pending'] += 1

    # Build calendar grid
    cal = calendar.monthcalendar(year, month)
    grid = []
    for week in cal:
        week_data = []
        for day in week:
            if day == 0:
                week_data.append(None)
            else:
                week_data.append({
                    'day': day,
                    'date': date(year, month, day),
                    'daily': daily.get(day),
                })
        grid.append(week_data)

    return render(request, 'finance/calendar.html', {
        'grid': grid,
        'year': year,
        'month': month,
        'month_name': calendar.month_name[month],
        'today': today,
        'prev': prev_month,
        'next': next_month_d,
        'day_names': ['Sun', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat'],
    })


@login_required
def calendar_day_detail(request, year, month, day):
    target_date = date(year, month, day)
    txns = Transaction.objects.filter(date=target_date).select_related('category', 'subcategory')
    total_income = txns.filter(transaction_type='Income').aggregate(s=Sum('amount'))['s'] or 0
    total_expense = txns.filter(transaction_type='Expense').aggregate(s=Sum('amount'))['s'] or 0
    return render(request, 'finance/calendar_day.html', {
        'transactions': txns,
        'target_date': target_date,
        'total_income': total_income,
        'total_expense': total_expense,
    })


# ─── Reports ───────────────────────────────────────────────────────────────────

@login_required
def reports(request):
    from datetime import date as ddate
    today = ddate.today()
    date_from = request.GET.get('date_from', today.replace(day=1).isoformat())
    date_to = request.GET.get('date_to', today.isoformat())

    qs = Transaction.objects.filter(date__gte=date_from, date__lte=date_to).select_related('category', 'subcategory').order_by('date')

    total_income = qs.filter(transaction_type='Income').aggregate(s=Sum('amount'))['s'] or Decimal('0')
    total_expense = qs.filter(transaction_type='Expense').aggregate(s=Sum('amount'))['s'] or Decimal('0')
    total_pending = qs.filter(status='Pending').aggregate(s=Sum('amount'))['s'] or Decimal('0')
    net_balance = total_income - total_expense

    if 'export_excel' in request.GET:
        return export_excel(qs, date_from, date_to, total_income, total_expense, total_pending, net_balance)
    if 'export_pdf' in request.GET:
        return export_pdf(qs, date_from, date_to, total_income, total_expense, total_pending, net_balance)

    return render(request, 'finance/reports.html', {
        'transactions': qs,
        'date_from': date_from,
        'date_to': date_to,
        'total_income': total_income,
        'total_expense': total_expense,
        'total_pending': total_pending,
        'net_balance': net_balance,
    })


def export_excel(qs, date_from, date_to, total_income, total_expense, total_pending, net_balance):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Transactions'

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1a56db', end_color='1a56db', fill_type='solid')
    income_fill = PatternFill(start_color='d1fae5', end_color='d1fae5', fill_type='solid')
    expense_fill = PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid')
    pending_fill = PatternFill(start_color='fef3c7', end_color='fef3c7', fill_type='solid')
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f'{settings.CHURCH_NAME} – Transaction Report'
    title_cell.font = Font(bold=True, size=14, color='1a56db')
    title_cell.alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:G2')
    date_cell = ws['A2']
    date_cell.value = f'Period: {date_from} to {date_to}'
    date_cell.alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Date', 'Type', 'Amount (₹)', 'Category', 'SubCategory', 'Notes', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin

    # Data
    for row_idx, txn in enumerate(qs, 5):
        row_data = [
            txn.date.strftime('%d-%m-%Y'),
            txn.transaction_type,
            float(txn.amount),
            txn.category.name,
            txn.subcategory.name if txn.subcategory else '',
            txn.notes,
            txn.status,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin
            if txn.transaction_type == 'Income':
                cell.fill = income_fill
            elif txn.status == 'Pending':
                cell.fill = pending_fill
            else:
                cell.fill = expense_fill
            if col == 3:
                cell.number_format = '₹#,##0.00'

    # Summary
    summary_row = ws.max_row + 2
    summary_data = [
        ('Total Income', float(total_income)),
        ('Total Expense', float(total_expense)),
        ('Total Pending', float(total_pending)),
        ('Net Balance', float(net_balance)),
    ]
    ws.cell(row=summary_row, column=1, value='SUMMARY').font = Font(bold=True, size=12)
    for i, (label, val) in enumerate(summary_data):
        r = summary_row + i + 1
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=val)
        c.number_format = '₹#,##0.00'
        if label == 'Total Income':
            c.font = Font(color='059669', bold=True)
        elif label == 'Total Expense':
            c.font = Font(color='dc2626', bold=True)
        elif label == 'Net Balance':
            c.font = Font(color='1a56db', bold=True)

    # Column widths
    col_widths = [14, 10, 14, 20, 20, 40, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="FaithLedger_Report_{date_from}_{date_to}.xlsx"'
    wb.save(response)
    return response


def export_pdf(qs, date_from, date_to, total_income, total_expense, total_pending, net_balance):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    import io

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), topMargin=1*cm, bottomMargin=1*cm,
                             leftMargin=1.5*cm, rightMargin=1.5*cm)

    styles = getSampleStyleSheet()
    blue = colors.HexColor('#1a56db')
    green = colors.HexColor('#059669')
    red = colors.HexColor('#dc2626')
    light_blue = colors.HexColor('#eff6ff')

    title_style = ParagraphStyle('Title', parent=styles['Title'], textColor=blue, fontSize=18, spaceAfter=4)
    sub_style = ParagraphStyle('Sub', parent=styles['Normal'], fontSize=10, textColor=colors.gray, spaceAfter=12, alignment=TA_CENTER)

    story = []
    story.append(Paragraph(settings.CHURCH_NAME, title_style))
    story.append(Paragraph('Transaction Report', ParagraphStyle('R', parent=styles['Normal'], fontSize=14, textColor=blue, alignment=TA_CENTER, spaceAfter=4)))
    story.append(Paragraph(f'Period: {date_from} to {date_to}', sub_style))
    story.append(Spacer(1, 0.3*cm))

    # Summary boxes
    summary_data = [
        ['Total Income', f'₹{total_income:,.2f}'],
        ['Total Expense', f'₹{total_expense:,.2f}'],
        ['Total Pending', f'₹{total_pending:,.2f}'],
        ['Net Balance', f'₹{net_balance:,.2f}'],
    ]
    sum_table = Table([summary_data[i:i+2] for i in range(0, 4, 2)] if False else [summary_data[i] for i in range(4)],
                      colWidths=[5*cm, 5*cm])
    sum_table = Table(
        [[Paragraph(f'<b>{k}</b>', styles['Normal']), Paragraph(v, ParagraphStyle('V', parent=styles['Normal'], alignment=1))]
         for k, v in summary_data],
        colWidths=[6*cm, 5*cm]
    )
    sum_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), light_blue),
        ('BOX', (0, 0), (-1, -1), 1, blue),
        ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.lightblue),
        ('PADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(sum_table)
    story.append(Spacer(1, 0.5*cm))

    # Transaction table
    table_data = [['Date', 'Type', 'Amount (₹)', 'Category', 'SubCategory', 'Notes', 'Status']]
    for txn in qs:
        table_data.append([
            txn.date.strftime('%d-%m-%Y'),
            txn.transaction_type,
            f'₹{txn.amount:,.2f}',
            txn.category.name,
            txn.subcategory.name if txn.subcategory else '-',
            txn.notes[:40] + ('...' if len(txn.notes) > 40 else '') if txn.notes else '-',
            txn.status,
        ])

    col_widths = [2.5*cm, 2*cm, 3*cm, 4*cm, 4*cm, 8*cm, 2.5*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)

    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), blue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (5, 1), (5, -1), 'LEFT'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('BOX', (0, 0), (-1, -1), 1, colors.lightgrey),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ('PADDING', (0, 0), (-1, -1), 4),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]

    for i, txn in enumerate(qs, 1):
        if txn.transaction_type == 'Income':
            table_style.append(('TEXTCOLOR', (1, i), (1, i), green))
            table_style.append(('TEXTCOLOR', (2, i), (2, i), green))
        else:
            table_style.append(('TEXTCOLOR', (1, i), (1, i), red))
            table_style.append(('TEXTCOLOR', (2, i), (2, i), red))
        if txn.status == 'Pending':
            table_style.append(('TEXTCOLOR', (6, i), (6, i), colors.HexColor('#d97706')))

    t.setStyle(TableStyle(table_style))
    story.append(t)

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="FaithLedger_Report_{date_from}_{date_to}.pdf"'
    return response


# ─── Analytics ─────────────────────────────────────────────────────────────────

@login_required
def analytics(request):
    today = date.today()
    year_start = today.replace(month=1, day=1)

    all_txns = Transaction.objects.all()

    # Highest income/expense category
    top_income_cat = all_txns.filter(transaction_type='Income').values('category__name').annotate(
        total=Sum('amount')).order_by('-total').first()
    top_expense_cat = all_txns.filter(transaction_type='Expense').values('category__name').annotate(
        total=Sum('amount')).order_by('-total').first()

    # Average monthly income (across all months with data)
    from django.db.models.functions import TruncMonth
    monthly_income = all_txns.filter(transaction_type='Income').annotate(
        month=TruncMonth('date')).values('month').annotate(total=Sum('amount'))
    avg_monthly_income = monthly_income.aggregate(avg=Avg('total'))['avg'] or Decimal('0')

    # Sunday offerings this year
    try:
        sunday_cat = Category.objects.get(name='Sunday Offerings', type='Income')
        sunday_offerings_year = all_txns.filter(
            category=sunday_cat, date__gte=year_start
        ).aggregate(s=Sum('amount'))['s'] or Decimal('0')
    except Category.DoesNotExist:
        sunday_offerings_year = Decimal('0')

    # Expense % by category
    total_expense = all_txns.filter(transaction_type='Expense').aggregate(s=Sum('amount'))['s'] or Decimal('1')
    expense_by_cat = all_txns.filter(transaction_type='Expense').values('category__name').annotate(
        total=Sum('amount')).order_by('-total')
    expense_pct = [
        {'name': e['category__name'], 'total': float(e['total']), 'pct': round(float(e['total']) / float(total_expense) * 100, 1)}
        for e in expense_by_cat
    ]

    # Monthly growth trend
    monthly_trend = []
    for i in range(11, -1, -1):
        d = today.replace(day=1) - timedelta(days=i * 28)
        ms = d.replace(day=1)
        me = (ms.replace(day=28) + timedelta(days=4)).replace(day=1)
        inc = all_txns.filter(date__gte=ms, date__lt=me, transaction_type='Income').aggregate(s=Sum('amount'))['s'] or 0
        exp = all_txns.filter(date__gte=ms, date__lt=me, transaction_type='Expense').aggregate(s=Sum('amount'))['s'] or 0
        monthly_trend.append({'month': ms.strftime('%b %y'), 'income': float(inc), 'expense': float(exp)})

    context = {
        'top_income_cat': top_income_cat,
        'top_expense_cat': top_expense_cat,
        'avg_monthly_income': avg_monthly_income,
        'sunday_offerings_year': sunday_offerings_year,
        'expense_pct': expense_pct,
        'monthly_trend': json.dumps(monthly_trend),
        'expense_by_cat_json': json.dumps([{'name': e['name'], 'total': e['total']} for e in expense_pct]),
    }
    return render(request, 'finance/analytics.html', context)


# ─── Categories ────────────────────────────────────────────────────────────────

@login_required
def category_list(request):
    categories = Category.objects.prefetch_related('subcategories').order_by('type', 'name')
    return render(request, 'finance/category_list.html', {'categories': categories})


@login_required
def category_add(request):
    form = CategoryForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Category added successfully!')
        return redirect('category_list')
    return render(request, 'finance/category_form.html', {'form': form, 'title': 'Add Category'})


@login_required
def category_edit(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    form = CategoryForm(request.POST or None, instance=cat)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Category updated successfully!')
        return redirect('category_list')
    return render(request, 'finance/category_form.html', {'form': form, 'title': 'Edit Category', 'cat': cat})


@login_required
def category_toggle(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    cat.is_active = not cat.is_active
    cat.save()
    status = 'enabled' if cat.is_active else 'disabled'
    messages.success(request, f'Category "{cat.name}" {status}.')
    return redirect('category_list')


@login_required
def subcategory_add(request):
    form = SubCategoryForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Sub-category added successfully!')
        return redirect('category_list')
    return render(request, 'finance/category_form.html', {'form': form, 'title': 'Add Sub-Category'})


@login_required
def subcategory_edit(request, pk):
    sub = get_object_or_404(SubCategory, pk=pk)
    form = SubCategoryForm(request.POST or None, instance=sub)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Sub-category updated!')
        return redirect('category_list')
    return render(request, 'finance/category_form.html', {'form': form, 'title': 'Edit Sub-Category'})


# ─── Register View ─────────────────────────────────────────────────────────────

def register_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    errors = {}
    form_data = {}

    if request.method == 'POST':
        form_data = request.POST
        first_name = request.POST.get('first_name', '').strip()
        last_name  = request.POST.get('last_name', '').strip()
        email      = request.POST.get('email', '').strip()
        username   = request.POST.get('username', '').strip()
        password1  = request.POST.get('password1', '')
        password2  = request.POST.get('password2', '')

        # Validation
        if not first_name:
            errors['first_name'] = 'First name is required.'
        if not last_name:
            errors['last_name'] = 'Last name is required.'
        if not email:
            errors['email'] = 'Email is required.'
        elif User.objects.filter(email=email).exists():
            errors['email'] = 'This email is already registered.'
        if not username:
            errors['username'] = 'Username is required.'
        elif User.objects.filter(username=username).exists():
            errors['username'] = 'This username is already taken.'
        elif len(username) < 3:
            errors['username'] = 'Username must be at least 3 characters.'
        if not password1:
            errors['password1'] = 'Password is required.'
        elif len(password1) < 8:
            errors['password1'] = 'Password must be at least 8 characters.'
        elif not any(c.isdigit() for c in password1):
            errors['password1'] = 'Password must contain at least one number.'
        elif not any(c in '!@#$%^&*(),.?":{}|<>_-+=[]\\/' for c in password1):
            errors['password1'] = 'Password must contain at least one special character.'
        if password1 and password2 and password1 != password2:
            errors['password2'] = 'Passwords do not match.'

        if not errors:
            user = User.objects.create_user(
                username=username,
                email=email,
                password=password1,
                first_name=first_name,
                last_name=last_name,
            )
            login(request, user)
            messages.success(request, f'Welcome to FaithLedger, {first_name}! Your account has been created.')
            return redirect('dashboard')

    features = ['Track income and expenses','Monthly reports and analytics','Calendar overview','Excel and PDF exports']
    return render(request, 'finance/register.html', {'errors': errors, 'form_data': form_data, 'features': features})
